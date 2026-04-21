#!/usr/bin/env python3
"""
Fluxo HTTP para Clinica Agil.

O script:
- autentica sem Selenium
- exporta o relatorio de vendas para o periodo informado
- exporta a planilha de pacientes
- persiste pacientes em SQLite com carga incremental por reexecucao
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import re
from db_util import connect as db_connect, sqlite3
import sys
import time
import unicodedata
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import requests
from env_config import default_env_file, env_int, load_env_file, runtime_root

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


BASE_URL = "https://app2.clinicaagil.com.br"

RUNTIME_ROOT = runtime_root()
DEFAULT_ENV_FILE = default_env_file()
DEFAULT_TIMEOUT = 60
DEFAULT_HTTP_MAX_ATTEMPTS = 4
DEFAULT_HTTP_BACKOFF_BASE_SECONDS = 2
DEFAULT_HTTP_BACKOFF_MAX_SECONDS = 20
DEFAULT_EMAIL = ""
DEFAULT_SENHA = ""
DEFAULT_DATA_VENDAS = "01/01/1900"
DEFAULT_DATA_PACIENTES = "01/01/1900"
DEFAULT_OUTPUT_DIR = RUNTIME_ROOT / "exports"
DEFAULT_DB_PATH = RUNTIME_ROOT / "mirella_pacientes.sqlite3"

LOGIN_CONFIG: Dict[str, Any] = {
    "base_url": BASE_URL,
    "endpoints": {
        "login": "/login",
        "relatorios": "/financeiro/relatorio/index",
        "vendas": "/financeiro/relatorio/vendas",
        "recebimentos": "/financeiro/relatorio/recebimentos",
        "pacientes": "/pacientes/exportar_xls",
    },
}

PATIENT_HEADER_TO_COLUMN = {
    "id": "patient_id",
    "nome": "nome",
    "data_nasc": "data_nasc",
    "telefone_1": "telefone_1",
    "telefone_2": "telefone_2",
    "telefone_3": "telefone_3",
    "matricula": "matricula",
    "convenio": "convenio",
    "sexo": "sexo",
    "etnia": "etnia",
    "responsaveis": "responsaveis",
    "nome_da_mae": "nome_mae",
    "cpf": "cpf",
    "identidade": "identidade",
    "cep": "cep",
    "endereco": "endereco",
    "e_mail": "email",
    "profissao": "profissao",
    "status": "status",
    "cidade": "cidade",
    "bairro": "bairro",
    "plano": "plano",
    "cpf_responsavel": "cpf_responsavel",
    "cns": "cns",
}

PATIENT_DATA_COLUMNS = [
    "nome",
    "data_nasc",
    "telefone_1",
    "telefone_2",
    "telefone_3",
    "matricula",
    "convenio",
    "sexo",
    "etnia",
    "responsaveis",
    "nome_mae",
    "cpf",
    "identidade",
    "cep",
    "endereco",
    "email",
    "profissao",
    "status",
    "cidade",
    "bairro",
    "plano",
    "cpf_responsavel",
    "cns",
]

FINANCIAL_SALES_HEADER_TO_COLUMN = {
    "competencia": "competencia",
    "interessado": "interessado",
    "categoria": "categoria",
    "subcategoria": "subcategoria",
    "observacoes": "observacoes",
    "vezes": "vezes",
    "total_bruto": "total_bruto",
    "total_liquido": "total_liquido",
    "tipo_de_pagamento": "tipo_pagamento",
    "cpf_cnpj_do_interessado": "cpf_cnpj_interessado",
    "profissional": "profissional",
}

FINANCIAL_RECEIPTS_HEADER_TO_COLUMN = {
    "interessado": "interessado",
    "vencimento": "vencimento",
    "pagamento": "pagamento",
    "procedimento_s_mensalidade": "procedimento_mensalidade",
    "profissional": "profissional",
    "valor_bruto": "valor_bruto",
    "valor_liquido": "valor_liquido",
}

FINANCIAL_SALES_COLUMNS = [
    "competencia",
    "interessado",
    "categoria",
    "subcategoria",
    "observacoes",
    "vezes",
    "total_bruto",
    "total_liquido",
    "tipo_pagamento",
    "cpf_cnpj_interessado",
    "profissional",
]

FINANCIAL_RECEIPTS_COLUMNS = [
    "interessado",
    "vencimento",
    "pagamento",
    "procedimento_mensalidade",
    "profissional",
    "valor_bruto",
    "valor_liquido",
]

PLACEHOLDER_STRINGS = {"'", "-", "--", "+55", "+55 "}

def _hoje_br() -> str:
    return datetime.now().strftime("%d/%m/%Y")


def _validar_data_br(valor: str) -> str:
    try:
        data = datetime.strptime(valor.strip(), "%d/%m/%Y")
    except ValueError as exc:
        raise ValueError(f"Data invalida '{valor}'. Use o formato dd/mm/aaaa.") from exc
    return data.strftime("%d/%m/%Y")


def _data_br_para_iso(valor: str) -> str:
    return datetime.strptime(_validar_data_br(valor), "%d/%m/%Y").strftime("%Y-%m-%d")


def _validar_periodo(data_de: str, data_ate: str, contexto: str) -> Tuple[str, str]:
    inicio = datetime.strptime(_validar_data_br(data_de), "%d/%m/%Y")
    fim = datetime.strptime(_validar_data_br(data_ate), "%d/%m/%Y")
    if inicio > fim:
        raise ValueError(f"Periodo invalido para {contexto}: data inicial maior que data final.")
    return inicio.strftime("%d/%m/%Y"), fim.strftime("%d/%m/%Y")


def _normalizar_header(valor: Any) -> str:
    texto = "" if valor is None else str(valor).strip().lower()
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"[^a-z0-9]+", "_", texto)
    return texto.strip("_")


def _normalizar_celula(valor: Any) -> Optional[str]:
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    texto = str(valor).strip()
    return texto or None


def _prompt_data(rotulo: str, valor_padrao: str, habilitado: bool) -> str:
    if not habilitado:
        return _validar_data_br(valor_padrao)
    while True:
        resposta = input(f"{rotulo} [{valor_padrao}]: ").strip()
        valor = resposta or valor_padrao
        try:
            return _validar_data_br(valor)
        except ValueError as exc:
            print(exc)


def _garantir_resposta_excel(resp: requests.Response, endpoint: str) -> None:
    if resp.status_code != 200:
        raise RuntimeError(f"Falha em {endpoint}: HTTP {resp.status_code}")
    content_type = (resp.headers.get("Content-Type") or "").lower()
    if "excel" in content_type or resp.content[:2] == b"PK":
        return
    raise RuntimeError(
        f"Resposta inesperada em {endpoint}. "
        f"Content-Type: {resp.headers.get('Content-Type', '<vazio>')}"
    )


def _celula_para_numero(valor: Any) -> float:
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        texto = valor.strip().replace("R$", "").strip()
        texto = texto.replace(".", "").replace(",", ".")
        try:
            return float(texto)
        except ValueError:
            return 0.0
    return 0.0


def _inserir_coluna_diferenca_vendas(caminho_xlsx: Path) -> None:
    if load_workbook is None:
        raise RuntimeError("openpyxl nao instalado. Instale com: pip install openpyxl")
    workbook = load_workbook(str(caminho_xlsx), read_only=False)
    worksheet = workbook.active
    worksheet.insert_cols(9)
    worksheet.cell(row=3, column=9).value = "Descontos"
    for row_idx in range(4, worksheet.max_row + 1):
        valor_g = _celula_para_numero(worksheet.cell(row=row_idx, column=7).value)
        valor_h = _celula_para_numero(worksheet.cell(row=row_idx, column=8).value)
        worksheet.cell(row=row_idx, column=9).value = valor_g - valor_h
    workbook.save(str(caminho_xlsx))
    workbook.close()


def _extrair_pacientes_do_xlsx(conteudo: bytes) -> List[Dict[str, Optional[str]]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl nao instalado. Instale com: pip install openpyxl")

    workbook = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    worksheet = workbook.active

    cabecalhos_norm: Optional[List[str]] = None
    registros: List[Dict[str, Optional[str]]] = []

    for row in worksheet.iter_rows(values_only=True):
        if not any(value not in (None, "") for value in row):
            continue

        if cabecalhos_norm is None:
            cabecalhos_norm = [_normalizar_header(value) for value in row]
            continue

        registro: Dict[str, Optional[str]] = {}
        for indice, header_norm in enumerate(cabecalhos_norm):
            if header_norm not in PATIENT_HEADER_TO_COLUMN:
                continue
            coluna = PATIENT_HEADER_TO_COLUMN[header_norm]
            registro[coluna] = _normalizar_celula(row[indice] if indice < len(row) else None)

        patient_id = registro.get("patient_id")
        if not patient_id:
            continue

        for coluna in PATIENT_DATA_COLUMNS:
            registro.setdefault(coluna, None)

        registros.append(registro)

    workbook.close()
    return registros


def _normalizar_nome_busca(valor: Any) -> Optional[str]:
    texto = _texto_ou_none(valor)
    if not texto:
        return None
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"[^a-z0-9]+", " ", texto.lower())
    texto = " ".join(texto.split())
    return texto or None


def _normalizar_documento_generico(valor: Any) -> Optional[str]:
    texto = _texto_ou_none(valor)
    if not texto:
        return None
    digits = re.sub(r"\D", "", texto)
    if not digits or set(digits) == {"0"}:
        return None
    return digits


def _numero_ou_none(valor: Any) -> Optional[float]:
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        texto = valor.strip().replace("R$", "").strip()
        if not texto:
            return None
        texto = texto.replace(".", "").replace(",", ".")
        try:
            return float(texto)
        except ValueError:
            return None
    return None


def _inteiro_ou_none(valor: Any) -> Optional[int]:
    numero = _numero_ou_none(valor)
    if numero is None:
        return None
    return int(numero)


def _extrair_linhas_financeiras_do_xlsx(
    conteudo: bytes,
    header_to_column: Dict[str, str],
    data_columns: Sequence[str],
    required_columns: Sequence[str],
) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl nao instalado. Instale com: pip install openpyxl")

    workbook = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    worksheet = workbook.active

    cabecalhos_norm: Optional[List[str]] = None
    header_row_number: Optional[int] = None
    registros: List[Dict[str, Any]] = []

    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if not any(value not in (None, "") for value in row):
            continue

        if cabecalhos_norm is None:
            candidato = [_normalizar_header(value) for value in row]
            colunas_detectadas = {header_to_column[h] for h in candidato if h in header_to_column}
            if all(coluna in colunas_detectadas for coluna in required_columns):
                cabecalhos_norm = candidato
                header_row_number = row_number
            continue

        registro: Dict[str, Any] = {"source_row_number": row_number}
        for indice, header_norm in enumerate(cabecalhos_norm):
            if header_norm not in header_to_column:
                continue
            coluna = header_to_column[header_norm]
            registro[coluna] = _normalizar_celula(row[indice] if indice < len(row) else None)

        if not any(registro.get(coluna) for coluna in data_columns):
            continue
        if not any(registro.get(coluna) for coluna in required_columns):
            continue

        for coluna in data_columns:
            registro.setdefault(coluna, None)
        registro["source_header_row_number"] = header_row_number
        registros.append(registro)

    workbook.close()
    return registros


def _extrair_vendas_do_xlsx(conteudo: bytes) -> List[Dict[str, Any]]:
    rows = _extrair_linhas_financeiras_do_xlsx(
        conteudo=conteudo,
        header_to_column=FINANCIAL_SALES_HEADER_TO_COLUMN,
        data_columns=FINANCIAL_SALES_COLUMNS,
        required_columns=("interessado", "total_bruto", "total_liquido"),
    )
    for row in rows:
        row["interessado_norm"] = _normalizar_nome_busca(row.get("interessado"))
        row["interessado_documento"] = _normalizar_documento_generico(row.get("cpf_cnpj_interessado"))
        row["vezes"] = _inteiro_ou_none(row.get("vezes"))
        row["total_bruto"] = _numero_ou_none(row.get("total_bruto"))
        row["total_liquido"] = _numero_ou_none(row.get("total_liquido"))
        bruto = row.get("total_bruto") or 0.0
        liquido = row.get("total_liquido") or 0.0
        row["descontos"] = bruto - liquido
    return rows


def _extrair_recebimentos_do_xlsx(conteudo: bytes) -> List[Dict[str, Any]]:
    rows = _extrair_linhas_financeiras_do_xlsx(
        conteudo=conteudo,
        header_to_column=FINANCIAL_RECEIPTS_HEADER_TO_COLUMN,
        data_columns=FINANCIAL_RECEIPTS_COLUMNS,
        required_columns=("interessado", "valor_bruto", "valor_liquido"),
    )
    for row in rows:
        row["interessado_norm"] = _normalizar_nome_busca(row.get("interessado"))
        row["valor_bruto"] = _numero_ou_none(row.get("valor_bruto"))
        row["valor_liquido"] = _numero_ou_none(row.get("valor_liquido"))
    return rows


def _texto_ou_none(valor: Any) -> Optional[str]:
    texto = _normalizar_celula(valor)
    if texto is None:
        return None
    if texto in PLACEHOLDER_STRINGS:
        return None
    return texto


def _normalizar_documento(valor: Any, tamanho: Optional[int] = None) -> Optional[str]:
    texto = _texto_ou_none(valor)
    if not texto:
        return None
    digits = re.sub(r"\D", "", texto)
    if not digits:
        return None
    if tamanho is not None and len(digits) != tamanho:
        return None
    if set(digits) == {"0"}:
        return None
    return digits


def _normalizar_email(valor: Any) -> Optional[str]:
    texto = _texto_ou_none(valor)
    if not texto:
        return None
    return texto.lower()


def _normalizar_telefone(valor: Any) -> Optional[str]:
    texto = _texto_ou_none(valor)
    if not texto:
        return None
    digits = re.sub(r"\D", "", texto)
    if not digits:
        return None
    significant = digits[2:] if digits.startswith("55") else digits
    if len(significant) < 10:
        return None
    if set(significant) == {"0"}:
        return None
    return digits


def _normalizar_data_br_ou_none(valor: Any) -> Optional[str]:
    texto = _texto_ou_none(valor)
    if not texto:
        return None
    try:
        return datetime.strptime(texto, "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return None


def _curar_registro_paciente(row: sqlite3.Row) -> Dict[str, Any]:
    patient_id = int(str(row["patient_id"]).strip())
    sexo = _texto_ou_none(row["sexo"])
    if sexo:
        sexo = sexo.upper()

    patient = {
        "patient_id": patient_id,
        "nome": _texto_ou_none(row["nome"]) or f"Paciente {patient_id}",
        "data_nascimento": _normalizar_data_br_ou_none(row["data_nasc"]),
        "sexo": sexo,
        "cpf": _normalizar_documento(row["cpf"], tamanho=11),
        "identidade": _texto_ou_none(row["identidade"]),
        "status": _texto_ou_none(row["status"]),
        "convenio": _texto_ou_none(row["convenio"]),
        "plano": _texto_ou_none(row["plano"]),
        "profissao": _texto_ou_none(row["profissao"]),
        "responsavel_nome": _texto_ou_none(row["responsaveis"]),
        "responsavel_cpf": _normalizar_documento(row["cpf_responsavel"], tamanho=11),
        "nome_mae": _texto_ou_none(row["nome_mae"]),
        "cns": _normalizar_documento(row["cns"]),
        "row_hash": row["row_hash"],
        "imported_at": row["imported_at"],
        "first_seen_at": row["first_seen_at"],
        "last_seen_at": row["last_seen_at"],
        "source_period_start": row["source_period_start"],
        "source_period_end": row["source_period_end"],
        "source_file_name": row["source_file_name"],
    }

    contacts: List[Dict[str, Any]] = []
    phone_contacts: List[Dict[str, Any]] = []
    for label, raw_value in (
        ("telefone_1", row["telefone_1"]),
        ("telefone_2", row["telefone_2"]),
        ("telefone_3", row["telefone_3"]),
    ):
        contact_value = _texto_ou_none(raw_value)
        contact_norm = _normalizar_telefone(raw_value)
        if contact_value and contact_norm:
            phone_contacts.append(
                {
                    "patient_id": patient_id,
                    "contact_type": "phone",
                    "contact_label": label,
                    "contact_value": contact_value,
                    "contact_value_norm": contact_norm,
                    "is_primary": 0,
                    "imported_at": row["imported_at"],
                }
            )

    if phone_contacts:
        phone_contacts[0]["is_primary"] = 1
        contacts.extend(phone_contacts)

    email_value = _texto_ou_none(row["email"])
    email_norm = _normalizar_email(row["email"])
    if email_value and email_norm:
        contacts.append(
            {
                "patient_id": patient_id,
                "contact_type": "email",
                "contact_label": "email_principal",
                "contact_value": email_value,
                "contact_value_norm": email_norm,
                "is_primary": 1,
                "imported_at": row["imported_at"],
            }
        )

    address = {
        "patient_id": patient_id,
        "cep": _normalizar_documento(row["cep"], tamanho=8),
        "endereco": _texto_ou_none(row["endereco"]),
        "bairro": _texto_ou_none(row["bairro"]),
        "cidade": _texto_ou_none(row["cidade"]),
        "imported_at": row["imported_at"],
    }

    return {
        "patient": patient,
        "contacts": contacts,
        "address": address,
    }


class SQLitePatientStore:
    def __init__(self, db_path: Path, logger: logging.Logger) -> None:
        self.db_path = Path(db_path)
        self.logger = logger
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = db_connect(self.db_path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys = ON")
        self._criar_schema()

    def close(self) -> None:
        self.conn.close()

    def _criar_schema(self) -> None:
        colunas_dados = ",\n                ".join(f"{coluna} TEXT" for coluna in PATIENT_DATA_COLUMNS)
        self.conn.executescript(
            f"""
            CREATE TABLE IF NOT EXISTS sync_state (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS patients_latest (
                patient_id TEXT PRIMARY KEY,
                row_hash TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                first_seen_at TEXT NOT NULL,
                last_seen_at TEXT NOT NULL,
                source_period_start TEXT NOT NULL,
                source_period_end TEXT NOT NULL,
                source_file_name TEXT,
                raw_payload_json TEXT NOT NULL,
                {colunas_dados}
            );

            CREATE TABLE IF NOT EXISTS patient_versions (
                patient_id TEXT NOT NULL,
                row_hash TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                source_period_start TEXT NOT NULL,
                source_period_end TEXT NOT NULL,
                source_file_name TEXT,
                raw_payload_json TEXT NOT NULL,
                PRIMARY KEY (patient_id, row_hash)
            );

            CREATE INDEX IF NOT EXISTS idx_patient_versions_imported_at
            ON patient_versions(imported_at);

            CREATE TABLE IF NOT EXISTS patient_import_runs (
                run_id INTEGER PRIMARY KEY AUTOINCREMENT,
                imported_at TEXT NOT NULL,
                source_period_start TEXT NOT NULL,
                source_period_end TEXT NOT NULL,
                source_file_name TEXT,
                total_rows INTEGER NOT NULL,
                inserted_rows INTEGER NOT NULL,
                updated_rows INTEGER NOT NULL,
                unchanged_rows INTEGER NOT NULL,
                versions_added INTEGER NOT NULL
            );

            CREATE TABLE IF NOT EXISTS patients (
                patient_id INTEGER PRIMARY KEY,
                nome TEXT NOT NULL,
                data_nascimento TEXT,
                sexo TEXT,
                cpf TEXT,
                identidade TEXT,
                status TEXT,
                convenio TEXT,
                plano TEXT,
                profissao TEXT,
                responsavel_nome TEXT,
                responsavel_cpf TEXT,
                nome_mae TEXT,
                cns TEXT,
                row_hash TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                first_seen_at TEXT NOT NULL,
                last_seen_at TEXT NOT NULL,
                source_period_start TEXT NOT NULL,
                source_period_end TEXT NOT NULL,
                source_file_name TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_patients_cpf ON patients(cpf);
            CREATE INDEX IF NOT EXISTS idx_patients_status ON patients(status);
            CREATE INDEX IF NOT EXISTS idx_patients_convenio ON patients(convenio);

            CREATE TABLE IF NOT EXISTS patient_contacts (
                contact_id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_id INTEGER NOT NULL,
                contact_type TEXT NOT NULL,
                contact_label TEXT NOT NULL,
                contact_value TEXT NOT NULL,
                contact_value_norm TEXT NOT NULL,
                is_primary INTEGER NOT NULL DEFAULT 0,
                imported_at TEXT NOT NULL,
                UNIQUE(patient_id, contact_type, contact_label),
                UNIQUE(patient_id, contact_type, contact_value_norm),
                FOREIGN KEY(patient_id) REFERENCES patients(patient_id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_patient_contacts_patient ON patient_contacts(patient_id);
            CREATE INDEX IF NOT EXISTS idx_patient_contacts_lookup
            ON patient_contacts(contact_type, contact_value_norm);

            CREATE TABLE IF NOT EXISTS patient_addresses (
                patient_id INTEGER PRIMARY KEY,
                cep TEXT,
                endereco TEXT,
                bairro TEXT,
                cidade TEXT,
                imported_at TEXT NOT NULL,
                FOREIGN KEY(patient_id) REFERENCES patients(patient_id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_patient_addresses_city ON patient_addresses(cidade);
            CREATE INDEX IF NOT EXISTS idx_patient_addresses_neighborhood ON patient_addresses(bairro);

            CREATE TABLE IF NOT EXISTS patient_financial_sales (
                sale_id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_period_start TEXT NOT NULL,
                source_period_end TEXT NOT NULL,
                source_file_name TEXT,
                source_row_number INTEGER NOT NULL,
                imported_at TEXT NOT NULL,
                matched_patient_id INTEGER,
                competencia TEXT,
                interessado TEXT,
                interessado_norm TEXT,
                interessado_documento TEXT,
                categoria TEXT,
                subcategoria TEXT,
                observacoes TEXT,
                vezes INTEGER,
                total_bruto REAL,
                total_liquido REAL,
                descontos REAL,
                tipo_pagamento TEXT,
                cpf_cnpj_interessado TEXT,
                profissional TEXT,
                row_hash TEXT NOT NULL,
                raw_payload_json TEXT NOT NULL,
                FOREIGN KEY(matched_patient_id) REFERENCES patients(patient_id) ON DELETE SET NULL
            );

            CREATE INDEX IF NOT EXISTS idx_patient_financial_sales_patient
            ON patient_financial_sales(matched_patient_id);

            CREATE INDEX IF NOT EXISTS idx_patient_financial_sales_doc
            ON patient_financial_sales(interessado_documento);

            CREATE INDEX IF NOT EXISTS idx_patient_financial_sales_period
            ON patient_financial_sales(source_period_start, source_period_end);

            CREATE TABLE IF NOT EXISTS patient_financial_receipts (
                receipt_id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_period_start TEXT NOT NULL,
                source_period_end TEXT NOT NULL,
                source_file_name TEXT,
                source_row_number INTEGER NOT NULL,
                imported_at TEXT NOT NULL,
                matched_patient_id INTEGER,
                interessado TEXT,
                interessado_norm TEXT,
                vencimento TEXT,
                pagamento TEXT,
                procedimento_mensalidade TEXT,
                profissional TEXT,
                valor_bruto REAL,
                valor_liquido REAL,
                row_hash TEXT NOT NULL,
                raw_payload_json TEXT NOT NULL,
                FOREIGN KEY(matched_patient_id) REFERENCES patients(patient_id) ON DELETE SET NULL
            );

            CREATE INDEX IF NOT EXISTS idx_patient_financial_receipts_patient
            ON patient_financial_receipts(matched_patient_id);

            CREATE INDEX IF NOT EXISTS idx_patient_financial_receipts_period
            ON patient_financial_receipts(source_period_start, source_period_end);

            DROP VIEW IF EXISTS vw_patients_complete_financial;
            DROP VIEW IF EXISTS vw_patients_complete;
            DROP VIEW IF EXISTS vw_patient_financial_summary;

            CREATE VIEW vw_patient_financial_summary AS
            WITH sales AS (
                SELECT
                    matched_patient_id AS patient_id,
                    COUNT(*) AS total_vendas_linhas,
                    COALESCE(SUM(total_bruto), 0) AS total_vendido_bruto,
                    COALESCE(SUM(total_liquido), 0) AS total_vendido_liquido,
                    COALESCE(SUM(descontos), 0) AS total_descontos
                FROM patient_financial_sales
                WHERE matched_patient_id IS NOT NULL
                GROUP BY matched_patient_id
            ),
            receipts AS (
                SELECT
                    matched_patient_id AS patient_id,
                    COUNT(*) AS total_recebimentos_linhas,
                    COALESCE(SUM(valor_bruto), 0) AS total_recebido_bruto,
                    COALESCE(SUM(valor_liquido), 0) AS total_recebido_liquido
                FROM patient_financial_receipts
                WHERE matched_patient_id IS NOT NULL
                GROUP BY matched_patient_id
            )
            SELECT
                p.patient_id,
                COALESCE(s.total_vendas_linhas, 0) AS total_vendas_linhas,
                COALESCE(s.total_vendido_bruto, 0) AS total_vendido_bruto,
                COALESCE(s.total_vendido_liquido, 0) AS total_vendido_liquido,
                COALESCE(s.total_descontos, 0) AS total_descontos,
                COALESCE(r.total_recebimentos_linhas, 0) AS total_recebimentos_linhas,
                COALESCE(r.total_recebido_bruto, 0) AS total_recebido_bruto,
                COALESCE(r.total_recebido_liquido, 0) AS total_recebido_liquido,
                COALESCE(s.total_vendido_liquido, 0) - COALESCE(r.total_recebido_liquido, 0) AS saldo_liquido_estimado
            FROM patients p
            LEFT JOIN sales s ON s.patient_id = p.patient_id
            LEFT JOIN receipts r ON r.patient_id = p.patient_id;

            CREATE VIEW vw_patients_complete AS
            SELECT
                p.patient_id,
                p.nome,
                p.data_nascimento,
                CASE
                    WHEN p.data_nascimento IS NOT NULL
                    THEN CAST((julianday('now') - julianday(p.data_nascimento)) / 365.2425 AS INTEGER)
                    ELSE NULL
                END AS idade,
                p.sexo,
                p.cpf,
                p.status,
                p.convenio,
                p.plano,
                p.profissao,
                a.cep,
                a.endereco,
                a.bairro,
                a.cidade,
                phone.contact_value AS telefone_principal,
                email.contact_value AS email_principal,
                p.imported_at,
                p.source_period_start,
                p.source_period_end
            FROM patients p
            LEFT JOIN patient_addresses a
                ON a.patient_id = p.patient_id
            LEFT JOIN patient_contacts phone
                ON phone.patient_id = p.patient_id
               AND phone.contact_type = 'phone'
               AND phone.is_primary = 1
            LEFT JOIN patient_contacts email
                ON email.patient_id = p.patient_id
               AND email.contact_type = 'email'
               AND email.is_primary = 1;

            CREATE VIEW vw_patients_complete_financial AS
            SELECT
                pc.*,
                fs.total_vendas_linhas,
                fs.total_vendido_bruto,
                fs.total_vendido_liquido,
                fs.total_descontos,
                fs.total_recebimentos_linhas,
                fs.total_recebido_bruto,
                fs.total_recebido_liquido,
                fs.saldo_liquido_estimado
            FROM vw_patients_complete pc
            LEFT JOIN vw_patient_financial_summary fs
                ON fs.patient_id = pc.patient_id;
            """
        )
        self.conn.commit()

    def get_state(self, key: str) -> Optional[str]:
        row = self.conn.execute("SELECT value FROM sync_state WHERE key = ?", (key,)).fetchone()
        return None if row is None else row["value"]

    def set_state(self, key: str, value: str) -> None:
        agora = datetime.now().isoformat(timespec="seconds")
        self.conn.execute(
            """
            INSERT INTO sync_state (key, value, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET
                value = excluded.value,
                updated_at = excluded.updated_at
            """,
            (key, value, agora),
        )

    def get_default_start_date(self) -> Optional[str]:
        return self.get_state("patients_last_sync_end_date")

    def _registrar_import_run(
        self,
        imported_at: str,
        source_period_start_iso: str,
        source_period_end_iso: str,
        source_file_name: str,
        resumo: Dict[str, int],
    ) -> None:
        self.conn.execute(
            """
            INSERT INTO patient_import_runs (
                imported_at,
                source_period_start,
                source_period_end,
                source_file_name,
                total_rows,
                inserted_rows,
                updated_rows,
                unchanged_rows,
                versions_added
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                imported_at,
                source_period_start_iso,
                source_period_end_iso,
                source_file_name,
                resumo["total"],
                resumo["inserted"],
                resumo["updated"],
                resumo["unchanged"],
                resumo["versions_added"],
            ),
        )

    def rebuild_curated_tables(self) -> Dict[str, int]:
        rows = self.conn.execute(
            "SELECT * FROM patients_latest ORDER BY CAST(patient_id AS INTEGER)"
        ).fetchall()

        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM patient_contacts")
        cursor.execute("DELETE FROM patient_addresses")
        cursor.execute("DELETE FROM patients")

        resumo = {
            "curated_patients": 0,
            "curated_contacts": 0,
            "curated_addresses": 0,
        }

        for row in rows:
            registro = _curar_registro_paciente(row)
            patient = registro["patient"]
            cursor.execute(
                """
                INSERT INTO patients (
                    patient_id,
                    nome,
                    data_nascimento,
                    sexo,
                    cpf,
                    identidade,
                    status,
                    convenio,
                    plano,
                    profissao,
                    responsavel_nome,
                    responsavel_cpf,
                    nome_mae,
                    cns,
                    row_hash,
                    imported_at,
                    first_seen_at,
                    last_seen_at,
                    source_period_start,
                    source_period_end,
                    source_file_name
                ) VALUES (
                    :patient_id,
                    :nome,
                    :data_nascimento,
                    :sexo,
                    :cpf,
                    :identidade,
                    :status,
                    :convenio,
                    :plano,
                    :profissao,
                    :responsavel_nome,
                    :responsavel_cpf,
                    :nome_mae,
                    :cns,
                    :row_hash,
                    :imported_at,
                    :first_seen_at,
                    :last_seen_at,
                    :source_period_start,
                    :source_period_end,
                    :source_file_name
                )
                """,
                patient,
            )
            resumo["curated_patients"] += 1

            address = registro["address"]
            if any(address[coluna] is not None for coluna in ("cep", "endereco", "bairro", "cidade")):
                cursor.execute(
                    """
                    INSERT INTO patient_addresses (
                        patient_id,
                        cep,
                        endereco,
                        bairro,
                        cidade,
                        imported_at
                    ) VALUES (
                        :patient_id,
                        :cep,
                        :endereco,
                        :bairro,
                        :cidade,
                        :imported_at
                    )
                    """,
                    address,
                )
                resumo["curated_addresses"] += 1

            for contact in registro["contacts"]:
                result = cursor.execute(
                    """
                    INSERT OR IGNORE INTO patient_contacts (
                        patient_id,
                        contact_type,
                        contact_label,
                        contact_value,
                        contact_value_norm,
                        is_primary,
                        imported_at
                    ) VALUES (
                        :patient_id,
                        :contact_type,
                        :contact_label,
                        :contact_value,
                        :contact_value_norm,
                        :is_primary,
                        :imported_at
                    )
                    """,
                    contact,
                )
                if result.rowcount:
                    resumo["curated_contacts"] += 1

        resumo.update(self.rematch_financial_rows(commit=False))
        return resumo

    def _patient_match_maps(self) -> Tuple[Dict[str, int], Dict[str, int]]:
        rows = self.conn.execute("SELECT patient_id, nome, cpf FROM patients").fetchall()
        by_document: Dict[str, int] = {}
        by_name: Dict[str, int] = {}
        duplicated_names: set[str] = set()

        for row in rows:
            patient_id = int(row["patient_id"])
            cpf = _normalizar_documento_generico(row["cpf"])
            if cpf:
                by_document[cpf] = patient_id

            nome_norm = _normalizar_nome_busca(row["nome"])
            if nome_norm:
                if nome_norm in by_name:
                    duplicated_names.add(nome_norm)
                else:
                    by_name[nome_norm] = patient_id

        for duplicated in duplicated_names:
            by_name.pop(duplicated, None)

        return by_document, by_name

    def _match_patient_id(
        self,
        by_document: Dict[str, int],
        by_name: Dict[str, int],
        documento: Optional[str],
        nome_norm: Optional[str],
    ) -> Optional[int]:
        if documento and documento in by_document:
            return by_document[documento]
        if nome_norm and nome_norm in by_name:
            return by_name[nome_norm]
        return None

    def rematch_financial_rows(self, commit: bool = True) -> Dict[str, int]:
        by_document, by_name = self._patient_match_maps()
        cursor = self.conn.cursor()
        resumo = {
            "financial_sales_rematched": 0,
            "financial_receipts_rematched": 0,
        }

        sales = cursor.execute(
            "SELECT sale_id, interessado_documento, interessado_norm FROM patient_financial_sales"
        ).fetchall()
        for row in sales:
            matched_patient_id = self._match_patient_id(
                by_document=by_document,
                by_name=by_name,
                documento=row["interessado_documento"],
                nome_norm=row["interessado_norm"],
            )
            cursor.execute(
                "UPDATE patient_financial_sales SET matched_patient_id = ? WHERE sale_id = ?",
                (matched_patient_id, row["sale_id"]),
            )
            if matched_patient_id is not None:
                resumo["financial_sales_rematched"] += 1

        receipts = cursor.execute(
            "SELECT receipt_id, interessado_norm FROM patient_financial_receipts"
        ).fetchall()
        for row in receipts:
            matched_patient_id = self._match_patient_id(
                by_document={},
                by_name=by_name,
                documento=None,
                nome_norm=row["interessado_norm"],
            )
            cursor.execute(
                "UPDATE patient_financial_receipts SET matched_patient_id = ? WHERE receipt_id = ?",
                (matched_patient_id, row["receipt_id"]),
            )
            if matched_patient_id is not None:
                resumo["financial_receipts_rematched"] += 1

        if commit:
            self.conn.commit()
        return resumo

    def replace_financial_sales(
        self,
        rows: Sequence[Dict[str, Any]],
        source_period_start: str,
        source_period_end: str,
        source_file_name: str,
    ) -> Dict[str, int]:
        imported_at = datetime.now().isoformat(timespec="seconds")
        source_period_start_iso = _data_br_para_iso(source_period_start)
        source_period_end_iso = _data_br_para_iso(source_period_end)
        by_document, by_name = self._patient_match_maps()

        cursor = self.conn.cursor()
        cursor.execute(
            """
            DELETE FROM patient_financial_sales
            WHERE source_period_start = ? AND source_period_end = ?
            """,
            (source_period_start_iso, source_period_end_iso),
        )

        resumo = {"total": 0, "matched": 0, "unmatched": 0}
        for row in rows:
            matched_patient_id = self._match_patient_id(
                by_document=by_document,
                by_name=by_name,
                documento=row.get("interessado_documento"),
                nome_norm=row.get("interessado_norm"),
            )
            if matched_patient_id is None:
                resumo["unmatched"] += 1
            else:
                resumo["matched"] += 1

            payload = {
                **row,
                "source_period_start": source_period_start_iso,
                "source_period_end": source_period_end_iso,
                "source_file_name": source_file_name,
                "imported_at": imported_at,
                "matched_patient_id": matched_patient_id,
            }
            payload_json = json.dumps(payload, ensure_ascii=False, sort_keys=True)
            row_hash = hashlib.sha256(payload_json.encode("utf-8")).hexdigest()
            cursor.execute(
                """
                INSERT INTO patient_financial_sales (
                    source_period_start,
                    source_period_end,
                    source_file_name,
                    source_row_number,
                    imported_at,
                    matched_patient_id,
                    competencia,
                    interessado,
                    interessado_norm,
                    interessado_documento,
                    categoria,
                    subcategoria,
                    observacoes,
                    vezes,
                    total_bruto,
                    total_liquido,
                    descontos,
                    tipo_pagamento,
                    cpf_cnpj_interessado,
                    profissional,
                    row_hash,
                    raw_payload_json
                ) VALUES (
                    :source_period_start,
                    :source_period_end,
                    :source_file_name,
                    :source_row_number,
                    :imported_at,
                    :matched_patient_id,
                    :competencia,
                    :interessado,
                    :interessado_norm,
                    :interessado_documento,
                    :categoria,
                    :subcategoria,
                    :observacoes,
                    :vezes,
                    :total_bruto,
                    :total_liquido,
                    :descontos,
                    :tipo_pagamento,
                    :cpf_cnpj_interessado,
                    :profissional,
                    :row_hash,
                    :raw_payload_json
                )
                """,
                {**payload, "row_hash": row_hash, "raw_payload_json": payload_json},
            )
            resumo["total"] += 1

        self.set_state("financial_sales_last_sync_start_date", source_period_start)
        self.set_state("financial_sales_last_sync_end_date", source_period_end)
        self.set_state("financial_sales_last_sync_at", imported_at)
        self.set_state("financial_sales_last_sync_rows", str(resumo["total"]))
        self.conn.commit()
        return resumo

    def replace_financial_receipts(
        self,
        rows: Sequence[Dict[str, Any]],
        source_period_start: str,
        source_period_end: str,
        source_file_name: str,
    ) -> Dict[str, int]:
        imported_at = datetime.now().isoformat(timespec="seconds")
        source_period_start_iso = _data_br_para_iso(source_period_start)
        source_period_end_iso = _data_br_para_iso(source_period_end)
        _, by_name = self._patient_match_maps()

        cursor = self.conn.cursor()
        cursor.execute(
            """
            DELETE FROM patient_financial_receipts
            WHERE source_period_start = ? AND source_period_end = ?
            """,
            (source_period_start_iso, source_period_end_iso),
        )

        resumo = {"total": 0, "matched": 0, "unmatched": 0}
        for row in rows:
            matched_patient_id = self._match_patient_id(
                by_document={},
                by_name=by_name,
                documento=None,
                nome_norm=row.get("interessado_norm"),
            )
            if matched_patient_id is None:
                resumo["unmatched"] += 1
            else:
                resumo["matched"] += 1

            payload = {
                **row,
                "source_period_start": source_period_start_iso,
                "source_period_end": source_period_end_iso,
                "source_file_name": source_file_name,
                "imported_at": imported_at,
                "matched_patient_id": matched_patient_id,
            }
            payload_json = json.dumps(payload, ensure_ascii=False, sort_keys=True)
            row_hash = hashlib.sha256(payload_json.encode("utf-8")).hexdigest()
            cursor.execute(
                """
                INSERT INTO patient_financial_receipts (
                    source_period_start,
                    source_period_end,
                    source_file_name,
                    source_row_number,
                    imported_at,
                    matched_patient_id,
                    interessado,
                    interessado_norm,
                    vencimento,
                    pagamento,
                    procedimento_mensalidade,
                    profissional,
                    valor_bruto,
                    valor_liquido,
                    row_hash,
                    raw_payload_json
                ) VALUES (
                    :source_period_start,
                    :source_period_end,
                    :source_file_name,
                    :source_row_number,
                    :imported_at,
                    :matched_patient_id,
                    :interessado,
                    :interessado_norm,
                    :vencimento,
                    :pagamento,
                    :procedimento_mensalidade,
                    :profissional,
                    :valor_bruto,
                    :valor_liquido,
                    :row_hash,
                    :raw_payload_json
                )
                """,
                {**payload, "row_hash": row_hash, "raw_payload_json": payload_json},
            )
            resumo["total"] += 1

        self.set_state("financial_receipts_last_sync_start_date", source_period_start)
        self.set_state("financial_receipts_last_sync_end_date", source_period_end)
        self.set_state("financial_receipts_last_sync_at", imported_at)
        self.set_state("financial_receipts_last_sync_rows", str(resumo["total"]))
        self.conn.commit()
        return resumo

    def upsert_patients(
        self,
        rows: Sequence[Dict[str, Optional[str]]],
        source_period_start: str,
        source_period_end: str,
        source_file_name: str,
    ) -> Dict[str, int]:
        imported_at = datetime.now().isoformat(timespec="seconds")
        source_period_start_iso = _data_br_para_iso(source_period_start)
        source_period_end_iso = _data_br_para_iso(source_period_end)

        insert_sql = f"""
            INSERT INTO patients_latest (
                patient_id,
                row_hash,
                imported_at,
                first_seen_at,
                last_seen_at,
                source_period_start,
                source_period_end,
                source_file_name,
                raw_payload_json,
                {", ".join(PATIENT_DATA_COLUMNS)}
            ) VALUES (
                :patient_id,
                :row_hash,
                :imported_at,
                :first_seen_at,
                :last_seen_at,
                :source_period_start,
                :source_period_end,
                :source_file_name,
                :raw_payload_json,
                {", ".join(f":{coluna}" for coluna in PATIENT_DATA_COLUMNS)}
            )
            ON CONFLICT(patient_id) DO UPDATE SET
                row_hash = excluded.row_hash,
                imported_at = excluded.imported_at,
                first_seen_at = patients_latest.first_seen_at,
                last_seen_at = excluded.last_seen_at,
                source_period_start = excluded.source_period_start,
                source_period_end = excluded.source_period_end,
                source_file_name = excluded.source_file_name,
                raw_payload_json = excluded.raw_payload_json,
                {", ".join(f"{coluna} = excluded.{coluna}" for coluna in PATIENT_DATA_COLUMNS)}
        """

        resumo = {
            "total": 0,
            "inserted": 0,
            "updated": 0,
            "unchanged": 0,
            "versions_added": 0,
        }

        cursor = self.conn.cursor()
        for row in rows:
            patient_id = row["patient_id"]
            payload_json = json.dumps(row, ensure_ascii=False, sort_keys=True)
            row_hash = hashlib.sha256(payload_json.encode("utf-8")).hexdigest()

            existente = cursor.execute(
                "SELECT row_hash FROM patients_latest WHERE patient_id = ?",
                (patient_id,),
            ).fetchone()

            if existente is None:
                resumo["inserted"] += 1
            elif existente["row_hash"] == row_hash:
                resumo["unchanged"] += 1
            else:
                resumo["updated"] += 1

            version_result = cursor.execute(
                """
                INSERT OR IGNORE INTO patient_versions (
                    patient_id,
                    row_hash,
                    imported_at,
                    source_period_start,
                    source_period_end,
                    source_file_name,
                    raw_payload_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    patient_id,
                    row_hash,
                    imported_at,
                    source_period_start_iso,
                    source_period_end_iso,
                    source_file_name,
                    payload_json,
                ),
            )
            if version_result.rowcount:
                resumo["versions_added"] += 1

            payload: Dict[str, Optional[str]] = {
                "patient_id": patient_id,
                "row_hash": row_hash,
                "imported_at": imported_at,
                "first_seen_at": imported_at,
                "last_seen_at": imported_at,
                "source_period_start": source_period_start_iso,
                "source_period_end": source_period_end_iso,
                "source_file_name": source_file_name,
                "raw_payload_json": payload_json,
            }
            for coluna in PATIENT_DATA_COLUMNS:
                payload[coluna] = row.get(coluna)

            cursor.execute(insert_sql, payload)
            resumo["total"] += 1

        self.set_state("patients_last_sync_end_date", source_period_end)
        self.set_state("patients_last_sync_start_date", source_period_start)
        self.set_state("patients_last_sync_at", imported_at)
        self.set_state("patients_last_sync_rows", str(resumo["total"]))
        self._registrar_import_run(
            imported_at=imported_at,
            source_period_start_iso=source_period_start_iso,
            source_period_end_iso=source_period_end_iso,
            source_file_name=source_file_name,
            resumo=resumo,
        )
        resumo.update(self.rebuild_curated_tables())
        self.set_state("patients_curated_count", str(resumo["curated_patients"]))
        self.set_state("patients_curated_contacts", str(resumo["curated_contacts"]))
        self.set_state("patients_curated_addresses", str(resumo["curated_addresses"]))
        self.conn.commit()
        return resumo


class ClinicaAgilHTTPExporter:
    def __init__(
        self,
        email: str,
        senha: str,
        timeout: int,
        output_dir: Path,
        logger: logging.Logger,
    ) -> None:
        self.email = email
        self.senha = senha
        self.timeout = timeout
        self.output_dir = Path(output_dir)
        self.logger = logger
        self.max_attempts = env_int("MIRELLA_HTTP_MAX_ATTEMPTS", DEFAULT_HTTP_MAX_ATTEMPTS)
        self.backoff_base_seconds = env_int("MIRELLA_HTTP_BACKOFF_BASE_SECONDS", DEFAULT_HTTP_BACKOFF_BASE_SECONDS)
        self.backoff_max_seconds = env_int("MIRELLA_HTTP_BACKOFF_MAX_SECONDS", DEFAULT_HTTP_BACKOFF_MAX_SECONDS)
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/145.0.0.0 Safari/537.36"
                )
            }
        )

    def _delay_for_attempt(self, attempt: int) -> int:
        delay = self.backoff_base_seconds * (2 ** max(0, attempt - 1))
        return min(self.backoff_max_seconds, delay)

    def _request_with_backoff(self, method: str, url: str, **kwargs: Any) -> requests.Response:
        last_error: Optional[Exception] = None
        last_response: Optional[requests.Response] = None
        for attempt in range(1, self.max_attempts + 1):
            try:
                response = self.session.request(method, url, timeout=self.timeout, **kwargs)
                if response.status_code == 429 and attempt < self.max_attempts:
                    delay = self._delay_for_attempt(attempt)
                    self.logger.warning(
                        "HTTP 429 em %s %s. Aguardando %ss antes da tentativa %s/%s.",
                        method,
                        url,
                        delay,
                        attempt + 1,
                        self.max_attempts,
                    )
                    time.sleep(delay)
                    last_response = response
                    continue
                return response
            except requests.RequestException as exc:
                last_error = exc
                if attempt >= self.max_attempts:
                    break
                delay = self._delay_for_attempt(attempt)
                self.logger.warning(
                    "Falha HTTP em %s %s: %s. Aguardando %ss antes da tentativa %s/%s.",
                    method,
                    url,
                    exc,
                    delay,
                    attempt + 1,
                    self.max_attempts,
                )
                time.sleep(delay)

        if last_error is not None:
            raise RuntimeError(f"Falha HTTP apos {self.max_attempts} tentativas: {method} {url} -> {last_error}") from last_error
        if last_response is not None:
            return last_response
        raise RuntimeError(f"Falha HTTP sem resposta apos {self.max_attempts} tentativas: {method} {url}")

    def autenticar(self, cookie_env: Optional[str] = None) -> None:
        if cookie_env:
            self.logger.info("Autenticacao: usando cookie de sessao informado.")
            self.session.headers["Cookie"] = cookie_env
            resposta = self._request_with_backoff(
                "GET",
                f"{BASE_URL}{LOGIN_CONFIG['endpoints']['relatorios']}",
            )
            if resposta.status_code == 200:
                return
            raise RuntimeError(f"Cookie de sessao invalido. HTTP {resposta.status_code}")

        self.logger.info("Autenticacao: fazendo login por HTTP.")
        resposta_login = self._request_with_backoff(
            "POST",
            f"{BASE_URL}{LOGIN_CONFIG['endpoints']['login']}",
            data={"identity": self.email, "password": self.senha},
        )
        if resposta_login.status_code not in (200, 302):
            if self.max_attempts > 1:
                delay = self._delay_for_attempt(self.max_attempts)
                self.logger.warning(
                    "Login rejeitado com HTTP %s. Aplicando atraso final de %ss para amortecer tentativas repetidas.",
                    resposta_login.status_code,
                    delay,
                )
                time.sleep(delay)
            raise RuntimeError(f"Falha no login. HTTP {resposta_login.status_code}")

        resposta_relatorios = self._request_with_backoff(
            "GET",
            f"{BASE_URL}{LOGIN_CONFIG['endpoints']['relatorios']}",
        )
        if resposta_relatorios.status_code != 200:
            raise RuntimeError(
                "Login retornou sem acesso a tela de relatorios. "
                f"HTTP {resposta_relatorios.status_code}"
            )
        self.logger.info("Autenticacao concluida com sucesso.")

    def _post_excel(self, endpoint: str, payload: Dict[str, str], nome: str) -> bytes:
        self.logger.info("Solicitando '%s' via HTTP...", nome)
        resposta = self._request_with_backoff("POST", f"{BASE_URL}{endpoint}", data=payload)
        _garantir_resposta_excel(resposta, endpoint)
        self.logger.info("Arquivo '%s' recebido (%.1f KB).", nome, len(resposta.content) / 1024.0)
        return resposta.content

    def exportar_vendas(self, data_de: str, data_ate: str) -> Dict[str, Path]:
        conteudo = self._post_excel(
            LOGIN_CONFIG["endpoints"]["vendas"],
            {
                "data_de": data_de,
                "data_ate": data_ate,
                "tipo": "xls",
            },
            "relatorio de vendas",
        )

        tag = f"{_data_br_para_iso(data_de)}_{_data_br_para_iso(data_ate)}"
        destino_dir = self.output_dir / "vendas"
        destino_dir.mkdir(parents=True, exist_ok=True)

        arquivo_profissionais = destino_dir / f"relatorio_vendas_profissionais_{tag}.xlsx"
        arquivo_geral = destino_dir / f"relatorio_vendas_geral_{tag}.xlsx"

        arquivo_profissionais.write_bytes(conteudo)
        arquivo_geral.write_bytes(conteudo)
        _inserir_coluna_diferenca_vendas(arquivo_geral)

        self.logger.info("Vendas salvas em: %s", arquivo_profissionais)
        self.logger.info("Vendas tratadas salvas em: %s", arquivo_geral)
        return {
            "profissionais": arquivo_profissionais,
            "geral": arquivo_geral,
        }

    def exportar_recebimentos(self, data_de: str, data_ate: str) -> Path:
        conteudo = self._post_excel(
            LOGIN_CONFIG["endpoints"]["recebimentos"],
            {
                "data_de": data_de,
                "data_ate": data_ate,
                "filtro_data": "vencimento",
                "tipo_plano_conta": "todos",
                "plano_conta": "",
                "tipo": "xls",
            },
            "relatorio de recebimentos",
        )

        tag = f"{_data_br_para_iso(data_de)}_{_data_br_para_iso(data_ate)}"
        destino_dir = self.output_dir / "vendas"
        destino_dir.mkdir(parents=True, exist_ok=True)
        arquivo = destino_dir / f"relatorio_recebimentos_{tag}.xlsx"
        arquivo.write_bytes(conteudo)
        self.logger.info("Recebimentos salvos em: %s", arquivo)
        return arquivo

    def exportar_pacientes(self, data_de: str, data_ate: str) -> Tuple[bytes, Path]:
        conteudo = self._post_excel(
            LOGIN_CONFIG["endpoints"]["pacientes"],
            {
                "data_de": data_de,
                "data_ate": data_ate,
            },
            "pacientes",
        )

        destino_dir = self.output_dir / "pacientes"
        destino_dir.mkdir(parents=True, exist_ok=True)
        arquivo_latest = destino_dir / "pacientes_list_latest.xlsx"
        arquivo_latest.write_bytes(conteudo)
        self.logger.info("Planilha mais recente de pacientes salva em: %s", arquivo_latest)
        return conteudo, arquivo_latest


def _resolver_periodo_vendas(args: argparse.Namespace) -> Optional[Tuple[str, str]]:
    if args.somente == "pacientes":
        return None

    pode_perguntar = not args.sem_input and sys.stdin.isatty()
    padrao_de = args.data_vendas_de or os.getenv("MIRELLA_DATA_VENDAS") or DEFAULT_DATA_VENDAS
    padrao_ate = args.data_vendas_ate or os.getenv("MIRELLA_DATA_VENDAS_ATE") or _hoje_br()

    data_de = _prompt_data("Data inicial do relatorio de vendas", padrao_de, pode_perguntar)
    data_ate = _prompt_data("Data final do relatorio de vendas", padrao_ate, pode_perguntar)
    return _validar_periodo(data_de, data_ate, "vendas")


def _resolver_periodo_pacientes(args: argparse.Namespace, store: SQLitePatientStore) -> Optional[Tuple[str, str]]:
    if args.somente == "vendas":
        return None

    if args.reprocessar_pacientes:
        padrao_de = (
            args.data_pacientes_de
            or os.getenv("MIRELLA_DATA_PACIENTES")
            or DEFAULT_DATA_PACIENTES
        )
    else:
        padrao_de = (
            args.data_pacientes_de
            or store.get_default_start_date()
            or os.getenv("MIRELLA_DATA_PACIENTES")
            or DEFAULT_DATA_PACIENTES
        )

    padrao_ate = args.data_pacientes_ate or os.getenv("MIRELLA_DATA_PACIENTES_ATE") or _hoje_br()
    return _validar_periodo(padrao_de, padrao_ate, "pacientes")


def _criar_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Exporta vendas e pacientes da Clinica Agil via HTTP e sincroniza pacientes em SQLite."
    )
    parser.add_argument("--email", default=os.getenv("MIRELLA_EMAIL", DEFAULT_EMAIL))
    parser.add_argument("--senha", default=os.getenv("MIRELLA_SENHA", DEFAULT_SENHA))
    parser.add_argument("--cookie", default=os.getenv("MIRELLA_COOKIE"))
    parser.add_argument("--timeout", type=int, default=env_int("MIRELLA_TIMEOUT", DEFAULT_TIMEOUT))
    parser.add_argument("--output-dir", default=str(Path(os.getenv("MIRELLA_OUTPUT_DIR", DEFAULT_OUTPUT_DIR))))
    parser.add_argument("--db-path", default=str(Path(os.getenv("MIRELLA_DB_PATH", DEFAULT_DB_PATH))))
    parser.add_argument("--somente", choices=("ambos", "vendas", "pacientes"), default="ambos")
    parser.add_argument("--sem-input", action="store_true", help="Nao pergunta datas no terminal.")
    parser.add_argument("--reprocessar-pacientes", action="store_true", help="Ignora a data da ultima sincronizacao.")
    parser.add_argument("--rebuild-curated", action="store_true", help="Reconstrui as tabelas curadas a partir de patients_latest sem chamar a API.")
    parser.add_argument("--data-vendas-de")
    parser.add_argument("--data-vendas-ate")
    parser.add_argument("--data-pacientes-de")
    parser.add_argument("--data-pacientes-ate")
    return parser


def main() -> None:
    load_env_file(DEFAULT_ENV_FILE)
    parser = _criar_parser()
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(message)s", datefmt="%H:%M:%S")
    logger = logging.getLogger(__name__)

    logger.info("============================================================")
    logger.info("Exportacao HTTP Clinica Agil")
    logger.info("Ambiente: %s", DEFAULT_ENV_FILE)
    logger.info("Saida: %s", args.output_dir)
    logger.info("SQLite: %s", args.db_path)
    logger.info("Modo: %s", args.somente)
    logger.info("============================================================")

    store = SQLitePatientStore(Path(args.db_path), logger)
    try:
        if args.rebuild_curated:
            resumo_curado = store.rebuild_curated_tables()
            store.set_state("patients_curated_count", str(resumo_curado["curated_patients"]))
            store.set_state("patients_curated_contacts", str(resumo_curado["curated_contacts"]))
            store.set_state("patients_curated_addresses", str(resumo_curado["curated_addresses"]))
            store.conn.commit()
            logger.info(
                "Tabelas curadas reconstruidas: pacientes=%s | contatos=%s | enderecos=%s",
                resumo_curado["curated_patients"],
                resumo_curado["curated_contacts"],
                resumo_curado["curated_addresses"],
            )
            return

        periodo_vendas = _resolver_periodo_vendas(args)
        periodo_pacientes = _resolver_periodo_pacientes(args, store)

        if periodo_vendas:
            logger.info("Periodo de vendas: %s ate %s", periodo_vendas[0], periodo_vendas[1])
        if periodo_pacientes:
            logger.info(
                "Periodo de pacientes: %s ate %s",
                periodo_pacientes[0],
                periodo_pacientes[1],
            )

        exporter = ClinicaAgilHTTPExporter(
            email=args.email,
            senha=args.senha,
            timeout=args.timeout,
            output_dir=Path(args.output_dir),
            logger=logger,
        )

        exporter.autenticar(cookie_env=args.cookie)

        if periodo_pacientes:
            conteudo_pacientes, arquivo_pacientes = exporter.exportar_pacientes(*periodo_pacientes)
            rows = _extrair_pacientes_do_xlsx(conteudo_pacientes)
            resumo = store.upsert_patients(
                rows=rows,
                source_period_start=periodo_pacientes[0],
                source_period_end=periodo_pacientes[1],
                source_file_name=arquivo_pacientes.name,
            )
            logger.info(
                "Pacientes sincronizados em SQLite: total=%s | novos=%s | atualizados=%s | sem_alteracao=%s | versoes_novas=%s | curated_patients=%s | curated_contacts=%s | curated_addresses=%s",
                resumo["total"],
                resumo["inserted"],
                resumo["updated"],
                resumo["unchanged"],
                resumo["versions_added"],
                resumo["curated_patients"],
                resumo["curated_contacts"],
                resumo["curated_addresses"],
            )

        if periodo_vendas:
            arquivos_vendas = exporter.exportar_vendas(*periodo_vendas)
            logger.info("Arquivos de vendas gerados: %s | %s", arquivos_vendas["profissionais"], arquivos_vendas["geral"])
            vendas_rows = _extrair_vendas_do_xlsx(arquivos_vendas["profissionais"].read_bytes())
            resumo_vendas = store.replace_financial_sales(
                rows=vendas_rows,
                source_period_start=periodo_vendas[0],
                source_period_end=periodo_vendas[1],
                source_file_name=arquivos_vendas["profissionais"].name,
            )
            logger.info(
                "Vendas sincronizadas em SQLite: total=%s | vinculadas=%s | sem_vinculo=%s",
                resumo_vendas["total"],
                resumo_vendas["matched"],
                resumo_vendas["unmatched"],
            )

            arquivo_recebimentos = exporter.exportar_recebimentos(*periodo_vendas)
            recebimentos_rows = _extrair_recebimentos_do_xlsx(arquivo_recebimentos.read_bytes())
            resumo_recebimentos = store.replace_financial_receipts(
                rows=recebimentos_rows,
                source_period_start=periodo_vendas[0],
                source_period_end=periodo_vendas[1],
                source_file_name=arquivo_recebimentos.name,
            )
            logger.info(
                "Recebimentos sincronizados em SQLite: total=%s | vinculados=%s | sem_vinculo=%s",
                resumo_recebimentos["total"],
                resumo_recebimentos["matched"],
                resumo_recebimentos["unmatched"],
            )

    finally:
        store.close()


def setup_driver(*_args: Any, **_kwargs: Any) -> None:
    raise RuntimeError("login.py agora e um fluxo HTTP. Use a execucao direta deste script.")


def realizar_login(*_args: Any, **_kwargs: Any) -> None:
    raise RuntimeError("login.py agora e um fluxo HTTP. Use a execucao direta deste script.")


def extrair_relatorio_vendas(*_args: Any, **_kwargs: Any) -> None:
    raise RuntimeError("login.py agora e um fluxo HTTP. Use a execucao direta deste script.")


def extrair_pacientes(*_args: Any, **_kwargs: Any) -> None:
    raise RuntimeError("login.py agora e um fluxo HTTP. Use a execucao direta deste script.")


if __name__ == "__main__":
    main()
